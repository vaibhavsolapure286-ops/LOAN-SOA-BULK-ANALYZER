"""
Excel Export Module for Loan SOA Bulk Analyzer
Creates formatted Excel reports with multiple sheets.
"""

import logging
from pathlib import Path
from typing import List, Dict
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from config import SHEET_NAMES, OUTPUT_EXCEL_FILE

# Setup logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)


class ExcelExporter:
    """Export analysis results to formatted Excel file."""

    def __init__(self, output_path: Path = OUTPUT_EXCEL_FILE):
        """
        Initialize the Excel exporter.

        Args:
            output_path: Path where Excel file will be saved
        """
        self.output_path = output_path
        self.sheets = {}

    @staticmethod
    def create_portfolio_summary(analyzed_loans: List[Dict]) -> pd.DataFrame:
        """
        Create portfolio summary dataframe.

        Args:
            analyzed_loans: List of analyzed loan dictionaries

        Returns:
            DataFrame with portfolio summary
        """
        summary_data = []

        for loan in analyzed_loans:
            summary_data.append(
                {
                    "Loan Number": loan.get("loan_number"),
                    "Customer Name": loan.get("customer_name"),
                    "Loan Amount (Rs.)": loan.get("loan_amount"),
                    "EMI (Rs.)": loan.get("emi_amount"),
                    "Current POS (Rs.)": loan.get("current_pos"),
                    "Total Bounces": loan.get("total_bounces"),
                    "Late EMI Count": loan.get("late_emi_count"),
                    "Risk Grade": loan.get("risk_grade"),
                }
            )

        df = pd.DataFrame(summary_data)

        # Add total row
        totals = {
            "Loan Number": "TOTAL",
            "Customer Name": "",
            "Loan Amount (Rs.)": df["Loan Amount (Rs.)"].sum(),
            "EMI (Rs.)": df["EMI (Rs.)"].sum(),
            "Current POS (Rs.)": df["Current POS (Rs.)"].sum(),
            "Total Bounces": df["Total Bounces"].sum(),
            "Late EMI Count": df["Late EMI Count"].sum(),
            "Risk Grade": "",
        }
        df = pd.concat([df, pd.DataFrame([totals])], ignore_index=True)

        return df

    @staticmethod
    def create_bounce_details(analyzed_loans: List[Dict]) -> pd.DataFrame:
        """
        Create bounce details dataframe.

        Args:
            analyzed_loans: List of analyzed loan dictionaries

        Returns:
            DataFrame with bounce details
        """
        bounce_data = []

        for loan in analyzed_loans:
            loan_number = loan.get("loan_number")
            customer_name = loan.get("customer_name")

            for bounce in loan.get("bounce_details", []):
                bounce_data.append(
                    {
                        "Loan Number": loan_number,
                        "Customer Name": customer_name,
                        "Bounce Date": bounce.get("date"),
                        "Description": bounce.get("description"),
                        "Amount (Rs.)": bounce.get("amount"),
                    }
                )

        return pd.DataFrame(bounce_data)

    @staticmethod
    def create_late_emi_details(analyzed_loans: List[Dict]) -> pd.DataFrame:
        """
        Create late EMI details dataframe.

        Args:
            analyzed_loans: List of analyzed loan dictionaries

        Returns:
            DataFrame with late EMI details
        """
        late_emi_data = []

        for loan in analyzed_loans:
            loan_number = loan.get("loan_number")
            customer_name = loan.get("customer_name")

            for late_emi in loan.get("late_emi_details", []):
                late_emi_data.append(
                    {
                        "Loan Number": loan_number,
                        "Customer Name": customer_name,
                        "Payment Date": late_emi.get("date"),
                        "Due Date": late_emi.get("due_date"),
                        "Days Late": late_emi.get("days_late"),
                        "Description": late_emi.get("description"),
                        "Amount (Rs.)": late_emi.get("amount"),
                    }
                )

        return pd.DataFrame(late_emi_data)

    @staticmethod
    def create_individual_loan_summary(analyzed_loans: List[Dict]) -> pd.DataFrame:
        """
        Create individual loan detailed summary.

        Args:
            analyzed_loans: List of analyzed loan dictionaries

        Returns:
            DataFrame with individual loan details
        """
        individual_data = []

        for loan in analyzed_loans:
            individual_data.append(
                {
                    "Loan Number": loan.get("loan_number"),
                    "Customer Name": loan.get("customer_name"),
                    "Loan Amount (Rs.)": loan.get("loan_amount"),
                    "EMI Amount (Rs.)": loan.get("emi_amount"),
                    "Current POS (Rs.)": loan.get("current_pos"),
                    "Total EMI Paid": loan.get("emi_count"),
                    "Total Bounces": loan.get("total_bounces"),
                    "Late EMI Count": loan.get("late_emi_count"),
                    "Risk Grade": loan.get("risk_grade"),
                    "PDF Source": loan.get("pdf_name"),
                }
            )

        return pd.DataFrame(individual_data)

    @staticmethod
    def format_currency_column(ws, column_letter: str):
        """
        Format a column as currency.

        Args:
            ws: Worksheet object
            column_letter: Column letter (e.g., 'B', 'C')
        """
        for cell in ws[column_letter]:
            if cell.row > 1:  # Skip header
                cell.number_format = "₹#,##0.00"

    @staticmethod
    def format_header_row(ws):
        """
        Format header row with styling.

        Args:
            ws: Worksheet object
        """
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = alignment
            cell.border = border

    @staticmethod
    def auto_adjust_columns(ws):
        """
        Automatically adjust column widths.

        Args:
            ws: Worksheet object
        """
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)

            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except TypeError:
                    pass

            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

    def export_to_excel(self, analyzed_loans: List[Dict]) -> bool:
        """
        Export analysis results to Excel file with multiple sheets.

        Args:
            analyzed_loans: List of analyzed loan dictionaries

        Returns:
            True if export successful, False otherwise
        """
        try:
            logger.info(f"Creating Excel export: {self.output_path}")

            # Create Excel writer
            with pd.ExcelWriter(self.output_path, engine="openpyxl") as writer:
                # Portfolio Summary
                portfolio_df = self.create_portfolio_summary(analyzed_loans)
                portfolio_df.to_excel(
                    writer,
                    sheet_name=SHEET_NAMES["portfolio_summary"],
                    index=False,
                )

                # Bounce Details
                bounce_df = self.create_bounce_details(analyzed_loans)
                bounce_df.to_excel(
                    writer,
                    sheet_name=SHEET_NAMES["bounce_details"],
                    index=False,
                )

                # Late EMI Details
                late_emi_df = self.create_late_emi_details(analyzed_loans)
                late_emi_df.to_excel(
                    writer,
                    sheet_name=SHEET_NAMES["late_emi_details"],
                    index=False,
                )

                # Individual Loan Summary
                individual_df = self.create_individual_loan_summary(analyzed_loans)
                individual_df.to_excel(
                    writer,
                    sheet_name=SHEET_NAMES["individual_loan_summary"],
                    index=False,
                )

            # Apply formatting
            self._apply_formatting()

            logger.info(
                f"Excel export successful: {self.output_path.absolute()}"
            )
            return True

        except Exception as e:
            logger.error(f"Error exporting to Excel: {e}")
            return False

    def _apply_formatting(self):
        """Apply formatting to Excel workbook."""
        try:
            wb = load_workbook(self.output_path)

            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]

                # Format header
                self.format_header_row(ws)

                # Auto-adjust columns
                self.auto_adjust_columns(ws)

                # Format currency columns
                for col in ws.iter_cols():
                    header = col[0].value
                    if header and "Rs." in str(header):
                        col_letter = get_column_letter(col[0].column)
                        self.format_currency_column(ws, col_letter)

            wb.save(self.output_path)
            logger.info("Excel formatting applied successfully")

        except Exception as e:
            logger.error(f"Error applying formatting: {e}")
